import MySQLdb
import requests
from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from bs4 import BeautifulSoup
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
import time
from openpyxl import load_workbook
import os
from difflib import SequenceMatcher
import re
#################절대 실행하지 말것 테이블 정보가 다 날라가는 수가 있음
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd     # pip install pands필요
import numpy as np
import yaml

with open('key.yaml') as f:
    conf = yaml.load(f)
    
options = webdriver.ChromeOptions()
options.add_experimental_option("prefs", {
  "download.default_directory": r"C:\Users\99rac\AppData\Local\Programs\Python\Python39",
  #r"C:\Users\99rac\AppData\Local\Programs\Python\Python39"
  "download.prompt_for_download": False,
  "download.directory_upgrade": True,
  "safebrowsing.enabled": True
})
#r"C:\Users\99rac\OneDrive\문서\Downloads\chromedriver_win32\chromedriver.exe"
chrome_driver=webdriver.Chrome(ChromeDriverManager().install(), chrome_options=options)
chrome_driver.implicitly_wait(5)



chrome_driver.get('https://clinicaltrials.gov/ct2/results?cond=&term=&cntry=KR&state=&city=&dist=')
chrome_driver.find_element_by_css_selector('#save-list-link').click()
chrome_driver.find_element_by_css_selector('#number-of-studies').click()
chrome_driver.find_element_by_css_selector('#number-of-studies > option:nth-child(4)').click()
chrome_driver.find_element_by_css_selector('#which-fields').click()
chrome_driver.find_element_by_css_selector('#which-fields > option:nth-child(2)').click()
chrome_driver.find_element_by_css_selector('#which-format').click()
chrome_driver.find_element_by_css_selector('#which-format > option:nth-child(4)').click()
chrome_driver.find_element_by_css_selector('#downloadForm > div:nth-child(2)').click()
chrome_driver.find_element_by_css_selector('#submit-download-list').click()
time.sleep(20)
# 해당이름의 csv파일을 읽어옴
r_csv = pd.read_csv("SearchResults.csv")

# 저장할 xlsx파일의 이름을 정함
save_xlsx = pd.ExcelWriter("SearchResults.xlsx")

r_csv.to_excel(save_xlsx, index = False) # xlsx 파일로 변환

save_xlsx.save() #xlsx 파일로 저장

###############################################################################clinicaltrial.gov에서 임상시험정보포함 엑셀 파일 다운로드 후 .xlsx파일로 저장

aws_user = conf['user']
aws_passwd = conf['passwd']
aws_host = conf['host']
aws_db = conf['db']

db=MySQLdb.connect(
    #User
    user=aws_user,
    #Test001
    passwd=aws_passwd,
    #localhost
    host=aws_host,
    #doctors,
    db=aws_db,
    charset="utf8"
)
curs = db.cursor() #Select Query #

curs.execute("DROP TABLE IF EXISTS ClinicalFinal")
curs.execute("CREATE TABLE ClinicalFinal (cid text, Trialname text,Status text, MeshTerm text, code text, StartDate text, CompletionDate text, url text, Name text,Hosp text,briefsummary text, detaileddescription text, cnt text)")


load_file_name='SearchResults.xlsx'

load_wb=load_workbook(os.path.join(os.getcwd(),load_file_name))
load_ws=load_wb['Sheet1']



get_TrialnameEngcells=load_ws['C2:C10001']
get_Statuscells=load_ws['E2:E10001']
get_MeshTermcells=load_ws['G2:G10001']
get_StartDatecells=load_ws['S2:S10001']
get_CompletionDatecells=load_ws['U2:U10001']
get_URLcells=load_ws['AA2:AA10001']

TrialnameEng=[]
Status=[]
MeshTerm=[]
StartDate=[]
PrimaryCompletionDate=[]
CompletionDate=[]
URL=[]

NameAndHospital=[]
briefsummary=[]
detaileddescription=[]


for i,rows in enumerate(get_TrialnameEngcells):
    for j,cell in enumerate(rows):
    #print(cell.value)
        TrialnameEng.append(cell.value)
    
for i,rows in enumerate(get_Statuscells):
    for j,cell in enumerate(rows):
    #print(cell.value)
        Status.append(cell.value)
    
for i,rows in enumerate(get_MeshTermcells):
    for j,cell in enumerate(rows):
    #print(cell.value)
        MeshTerm.append(cell.value)
    
for i,rows in enumerate(get_StartDatecells):
    for j,cell in enumerate(rows):
        #print(cell.value)
        StartDate.append(cell.value)


for i,rows in enumerate(get_CompletionDatecells):
    for j,cell in enumerate(rows):
    #print(cell.value)
        CompletionDate.append(cell.value)
    
for i,rows in enumerate(get_URLcells):
    for j,cell in enumerate(rows):
    #print(cell.value)
        URL.append(cell.value)

#print(TrialnameEng[1])
#print(Status[1])
#print(MeshTerm[1])
#print(StartDate[1])
#print(CompletionDate[1])
"""
for i in range(len(Status)):
sql="insert into Trials (Trialname,Status,MeshTerm, StartDate, CompletionDate, url) values(%s,%s,%s,%s,%s,%s)"
curs.execute(sql,(TrialnameEng[i],Status[i],MeshTerm[i], StartDate[i], CompletionDate[i], URL[i]))
db.commit()
"""



#chrome_driver = webdriver.Chrome(r"C:\Users\99rac\OneDrive\문서\Downloads\chromedriver_win32\chromedriver.exe")
#chrome_driver.implicitly_wait(5)


for u in range(len(URL)):
    """
    if u==10:
        break
    """
    flag=1
    try:
        if "http" in URL[u]:
            flag=0
            try:
                chrome_driver.get(URL[u])
            except TimeoutException:
                NameAndHospital.append("NULL")
            
            try:
                Text = chrome_driver.find_element_by_css_selector('#responsibleparty').text
            except (AttributeError,NoSuchElementException,TimeoutException):
                Text="a"
            if "," not in Text:
                NameAndHospital.append("NULL")
        
            elif "ltd" in Text:
                NameAndHospital.append("NULL")
        
            elif "Ltd" in Text:
                NameAndHospital.append("NULL")
            elif "Inc" in Text:
                NameAndHospital.append("NULL")
            else:
                NameAndHospital.append(Text)
            time.sleep(5)


            try:
                briefsummary.append(chrome_driver.find_element_by_css_selector('#tab-body > div > div:nth-child(1) > div:nth-child(2) > div.ct-body3.tr-indent2').text)
            #print(briefsummary)
            except (NoSuchElementException,TimeoutException):
                briefsummary.append('NULL')

            try:
                detaileddescription.append(chrome_driver.find_element_by_css_selector('#COLLAPSE-DetailedDesc > div.ct-body3.tr-indent2').text)
            #print(detaileddescription)
            except (NoSuchElementException,TimeoutException):
                detaileddescription.append('NULL')
        
        if flag==1:
            NameAndHospital.append("NULL")

    except TypeError:
        NameAndHospital.append("NULL")





#print(NameAndHospital)
#################################################################################### .xlsx 파일 이용, 해당 엑셀에 없는 정보 추가 크롤링 위해서 url링크로 접속 후 크롤링
"""
load_file_name2='심장질환+희귀질환표시_final.xlsx'

load_wb2=load_workbook(os.path.join(os.getcwd(),load_file_name2))
load_ws2=load_wb2['Sheet1']
get_Kcells2=load_ws2['D2:D1460']
get_Ecells=load_ws2['E2:E1460']
get_Ccells=load_ws2['A2:A1460']
get_fcells=load_ws2['F2:F1460']
Korean=[]
English=[]
code=[]
filter=[]
for i,rows in enumerate(get_Kcells2):
    for j,cell in enumerate(rows):
    #print(cell.value)
        Korean.append(cell.value)
for i,rows in enumerate(get_Ecells):
    for j,cell in enumerate(rows):
    #print(cell.value)
        English.append(cell.value)
for i, rows in enumerate(get_Ccells):
    for j, cell in enumerate(rows):
   #print(cell.value)
       code.append(cell.value)
for i, rows in enumerate(get_fcells):
    for j, cell in enumerate(rows):
    #print(cell.value)
        filter.append(cell.value)

curs.execute("DROP TABLE IF EXISTS dtable_final")

sql= "insert into dtable_final(korean,english,code,filter) values(%s,%s,%s,%s)"
curs.execute('CREATE TABLE dtable_final(Korean text, English text,code text,filter text)')


for i in range(len(Korean)):
    curs.execute(sql,(Korean[i], English[i],code[i],filter[i]))
db.commit()
"""

sql="select * from clinicaldisease"
md=pd.read_sql_query(sql,db)
################################################################################### 질병코드매칭에 쓰일 질병테이블 가져오기

fakeMeshlists=[]
Meshlists=[]
temp=[]
Codelists=[]
Codelist=[]
i=0
flag=0
flag2=0
for m in range(len(MeshTerm)):#len(MeshTerm)

    
    
    if NameAndHospital[m]=='NULL':
        Codelist.append('NULL')
        continue
    try:
        fakeMeshlists=re.split('[|,]',MeshTerm[m])
        i+=1
    
    except TypeError:
        flag+=1
        Codelist.append('NULL')
        continue

    for f in fakeMeshlists:
        temp=f.split(' and ')
        for t in temp:
            Meshlists.append(t)

    for M in Meshlists:
        for e in range(len(md['disease'])):
            if M in md['disease']:
                Codelists.append(md['code'][e])
                flag2=1

    if flag2==0:
        Codelist.append('NULL')
    if flag2==1:
        result=list(set(Codelists))
        Codelist.append(",".join(result))
        flag2=0

    Codelists=[]
    Meshlists=[]


print(i+flag)

##############################################################################################임상시험질병과 질병코드 매칭

Name=[]
Hosp=[]
Name4Kor=[]
Hosp4Kor=[]
deletelist=['Ph.D','PhD.','PhD','Seoul', 'Korea','MD,PhD','MD phD','OMD','MD. PhD.','M.D','MHS','phD','MMS','PhD','Ph.D','Ph.D','PhD.','M.D.','MD','KMD','MD PhD','M.D.-Ph.D.','Ph.D.','Dr.','Noh (PhD)','Principal Investigator']
Medlist=['Center','Hospital','Foundation','University']
flag2=0
i=0
for g in range(len(NameAndHospital)):

    

    NameandHosp=NameAndHospital[g].split(', ')

    if len(NameandHosp)==2:
    #i+=1
        Name.append(NameandHosp[0])
        Hosp.append(NameandHosp[1])
    #print(NameandHosp)
    elif len(NameandHosp)>=3:
        for N in NameandHosp:
            for d in deletelist:
                if N==d:
                    NameandHosp.remove(N)
                if 'PhD' in NameandHosp:
                    NameandHosp.remove('PhD')
                if 'Ph.D' in NameandHosp:
                    NameandHosp.remove('Ph.D')
                if 'PhD.' in NameandHosp:
                    NameandHosp.remove('PhD.')
                if 'phD' in NameandHosp:
                    NameandHosp.remove('phD')
                if 'Ph.D.' in NameandHosp:
                    NameandHosp.remove('Ph.D.')
                if 'MMS' in NameandHosp:
                    NameandHosp.remove('MMS')
                if 'MHS' in NameandHosp:
                    NameandHosp.remove('MHS')
                if 'Korea' in NameandHosp:
                    NameandHosp.remove('Korea')
                if 'Chung-Ang University College of Medicine' in NameandHosp:
                    NameandHosp.remove('Chung-Ang University College of Medicine')
        if len(NameandHosp)>=3:
            Name.append(NameandHosp[0]+' '+NameandHosp[1])
            Hosp.append(NameandHosp[2])
        #print(Hosp)

#N=Name[0]
#H=Hosp[0]
#Name4Kor=get_translate(Name)
#Hosp4Kor=get_translate(Hosp)


    sql="insert into ClinicalFinal(cid, Trialname, Status, MeshTerm, code, StartDate, CompletionDate, url, Name, Hosp,briefsummary, detaileddescription) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"

#curs.execute("CREATE TABLE ClinicalFinal (Trialname text,Status text, MeshTerm text, StartDate text, CompletionDate text, url text, Name text,Hosp text, Name4Kor text, Hosp4Kor text, cnt text)")

    try:
        curs.execute(sql,(g,TrialnameEng[g],Status[g],MeshTerm[g],Codelist[g], StartDate[g], CompletionDate[g], URL[g],Name,Hosp,briefsummary[g],detaileddescription[g]))
    except MySQLdb._exceptions.ProgrammingError:
        Name='NULL'
        Hosp='NULL'
        Name4Kor='NULL'
        Hosp4Kor='NULL'
        curs.execute(sql,(g,TrialnameEng[g],Status[g],MeshTerm[g],Codelist[g], StartDate[g], CompletionDate[g], URL[g],Name,Hosp,briefsummary[g],detaileddescription[g]))
    except IndexError:
        pass
    db.commit()

    Name4Kor=[]
    Hosp4Kor=[]

    Name=[]
    Hosp=[]
    db.commit()
    
